import os
import sys
import pandas as pd
import sqlite3
import unicodedata
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xml.etree.ElementTree as ET
import shutil
from pathlib import Path

DATA_DIR = "/app/data"
RESULT_DIR = "/result"
BACKUP_DIR = "/backup"
DB_FILE = os.path.join(DATA_DIR, "data.db")

def remove_diacritics(text):
    """Odstraní diakritiku z textu."""
    if not isinstance(text, str):
        return ""
    text = unicodedata.normalize('NFKD', text)
    return ''.join([c for c in text if not unicodedata.combining(c)])

def normalize_name(name):
    """Odstranění diakritiky a převod na malá písmena bez přebytečných mezer (pro porovnání)."""
    name = remove_diacritics(name)
    return ' '.join(name.lower().split())

def normalize_name_for_excel(name):
    """Odstraní diakritiku, každé slovo s velkým písmenem (pro výstup do Excelu)."""
    name = remove_diacritics(name)
    name = ' '.join(name.lower().split())
    return ' '.join([w.capitalize() for w in name.split()])

def create_backup_dir(excel_file, xml_file, db_file):
    """Vytvoří záložní složku a zkopíruje do ní soubory."""
    today = datetime.now(ZoneInfo("Europe/Prague")).strftime('%-d.%-m.%Y')
    day_dir = os.path.join(BACKUP_DIR, today)
    os.makedirs(day_dir, exist_ok=True)

    existing = [int(d) for d in os.listdir(day_dir) if d.isdigit()]
    next_num = max(existing, default=0) + 1
    backup_dir = os.path.join(day_dir, str(next_num))
    os.makedirs(backup_dir)

    print(f"\nVytvářím zálohu v: {backup_dir}")
    try:
        shutil.copy2(excel_file, os.path.join(backup_dir, Path(excel_file).name))
        print(f"Excel soubor zálohován: {Path(excel_file).name}")

        shutil.copy2(xml_file, os.path.join(backup_dir, Path(xml_file).name))
        print(f"XML soubor zálohován: {Path(xml_file).name}")

        shutil.copy2(db_file, os.path.join(backup_dir, Path(db_file).name))
        print(f"Databáze zálohována: {Path(db_file).name}")
    except Exception as e:
        print(f"Chyba při zálohování: {str(e)}")
        return None

    return backup_dir

def excel_to_sqlite(excel_path, db_path):
    print(f"Načítám data z: {excel_path}")
    df = pd.read_excel(excel_path)
    print(f"Ukládám data do databáze: {db_path} (tabulka 'excel_data')")
    conn = sqlite3.connect(db_path)
    df.to_sql("excel_data", conn, if_exists="replace", index=False)
    conn.close()
    print("Hotovo! Data z Excelu jsou v tabulce 'excel_data'.")

def xml_to_sqlite(xml_path, db_path):
    print(f"Načítám data z: {xml_path}")
    ns = {'vzp': 'http://xmlns.gemsystem.cz/RSZP/ZamestnanciPojisteniUVZP/1.0'}
    with open(xml_path, encoding="iso-8859-2"):
        tree = ET.parse(xml_path)
    root = tree.getroot()
    zamestnanci = []
    for zam in root.findall('.//vzp:zamestnanec', ns):
        cislo = zam.find('vzp:cisloPojistence', ns)
        jmeno = zam.find('vzp:jmeno', ns)
        prijmeni = zam.find('vzp:prijmeni', ns)
        od = zam.find('vzp:zamestnaniOd', ns)
        do = zam.find('vzp:zamestnaniDo', ns)
        kod = zam.find('vzp:kodPojisteni', ns)
        stav = zam.find('vzp:stavy/vzp:stav', ns)
        stat_kat = zam.find('vzp:statniKategorie/vzp:statniKategorieZamestnanec/vzp:kod', ns)
        stat_od = zam.find('vzp:statniKategorie/vzp:statniKategorieZamestnanec/vzp:statniKategorieOd', ns)
        stat_do = zam.find('vzp:statniKategorie/vzp:statniKategorieZamestnanec/vzp:statniKategorieDo', ns)
        zamestnanci.append({
            'cisloPojistence': cislo.text if cislo is not None else "",
            'jmeno': jmeno.text if jmeno is not None else "",
            'prijmeni': prijmeni.text if prijmeni is not None else "",
            'zamestnaniOd': od.text if od is not None else "",
            'zamestnaniDo': do.text if do is not None else "",
            'kodPojisteni': kod.text if kod is not None else "",
            'stav': stav.text if stav is not None else "",
            'statniKategorieOd': stat_od.text if stat_od is not None else "",
            'statniKategorieDo': stat_do.text if stat_do is not None else "",
        })
    df = pd.DataFrame(zamestnanci)
    print(f"Načteno {len(df)} záznamů ze souboru {xml_path}")
    if df.empty:
        print("Chyba: DataFrame z XML je prázdný! Pravděpodobně nebyly nalezeny žádné záznamy v XML.")
        print("Zkontrolujte správnost namespace a strukturu XML.")
        print("Sloupce DataFrame:", df.columns)
        sys.exit(1)
    print(f"Ukládám data do databáze: {db_path} (tabulka 'xml_data')")
    conn = sqlite3.connect(db_path)
    df.to_sql("xml_data", conn, if_exists="replace", index=False)
    conn.close()
    print("Hotovo! Data z XML jsou v tabulce 'xml_data'.")

def compare_and_export():
    print("Porovnávám data podle rod_cislo/cisloPojistence...")
    with open('/app/EXCEL_sql_s_diakritikou.sql', encoding='utf-8') as f:
        sql_excel = f.read()
    with open('/app/XML_sql_s_diakritikou.sql', encoding='utf-8') as f:
        sql_xml = f.read()

    conn = sqlite3.connect(DB_FILE)
    df_excel = pd.read_sql_query(sql_excel, conn)
    df_xml = pd.read_sql_query(sql_xml, conn)
    conn.close()

    # Normalizace klíčů a jmen
    df_excel['rod_cislo_norm'] = df_excel.iloc[:,0].astype(str).str.zfill(10)
    df_excel['jmeno_prijmeni_norm'] = df_excel.iloc[:,1].apply(normalize_name)
    df_xml['cisloPojistence_norm'] = df_xml.iloc[:,0].astype(str).str.zfill(10)
    df_xml['jmeno_prijmeni_norm'] = df_xml.iloc[:,1].apply(normalize_name)

    # Merge podle klíče (Excel LEFT JOIN XML)
    merged = pd.merge(
        df_excel,
        df_xml,
        left_on='rod_cislo_norm',
        right_on='cisloPojistence_norm',
        how='left',
        suffixes=('', '_xml')
    )

    vysledek = []
    for i, row in merged.iterrows():
        if pd.isna(row['cisloPojistence_norm']):
            vysledek.append('NEPRAVDA')
        else:
            if row['jmeno_prijmeni_norm'] == row['jmeno_prijmeni_norm_xml']:
                vysledek.append('PRAVDA')
            else:
                vysledek.append('NEPRAVDA')

    # Výstupní jména bez diakritiky a s velkým počátečním písmenem
    jmeno_excel = merged.iloc[:,1].apply(normalize_name_for_excel)
    jmeno_xml = merged['jmeno_prijmeni_xml'].apply(normalize_name_for_excel) if 'jmeno_prijmeni_xml' in merged else pd.Series([""]*len(merged))

    out_main = pd.DataFrame({
        'rod_cislo': merged.iloc[:,0].astype(str).str.zfill(10),
        'jmeno_prijmeni': jmeno_excel,
        'cisloPojistence': merged['cisloPojistence'].astype(str).str.zfill(10).where(merged['cisloPojistence'].notna(), None),
        'jmeno_prijmeni.1': jmeno_xml,
        'vysledek': vysledek
    })

    # Najít záznamy z XML, které nejsou v Excelu (XML ONLY)
    xml_only = df_xml[~df_xml['cisloPojistence_norm'].isin(df_excel['rod_cislo_norm'])].copy()
    xml_only['cisloPojistence'] = xml_only.iloc[:,0].astype(str).str.zfill(10)
    xml_only['jmeno_prijmeni'] = xml_only.iloc[:,1].apply(normalize_name_for_excel)
    out_xml_only = xml_only[['cisloPojistence', 'jmeno_prijmeni']].reset_index(drop=True)

    # Zarovnat počet řádků
    max_len = max(len(out_main), len(out_xml_only))
    out_main = out_main.reindex(range(max_len)).reset_index(drop=True)
    out_xml_only = out_xml_only.reindex(range(max_len)).reset_index(drop=True)

    # Složit výsledný DataFrame
    out = pd.concat([out_main, pd.DataFrame({'': ['']*max_len, ' ': ['']*max_len}), out_xml_only], axis=1)
    out.columns = [
        'rod_cislo', 'jmeno_prijmeni', 'cisloPojistence', 'jmeno_prijmeni.1', 'vysledek',
        '', ' ',
        'cisloPojistence_navic', 'jmeno_prijmeni_navic'
    ]

    # Uložení do Excelu
    if not os.path.exists(RESULT_DIR):
        os.makedirs(RESULT_DIR)
    now = datetime.now(ZoneInfo("Europe/Prague")).strftime('%d.%m.%Y_%H-%M')
    excel_path = os.path.join(RESULT_DIR, f'result_{now}.xlsx')
    out.to_excel(excel_path, index=False)

    # Barvení v Excelu
    wb = load_workbook(excel_path)
    ws = wb.active

    green = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
    red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    orange = PatternFill(start_color="FFFFC000", end_color="FFFFC000", fill_type="solid")

    for i, row in out.iterrows():
        excel_row = i + 2
        ws.cell(row=excel_row, column=1).number_format = '@'
        ws.cell(row=excel_row, column=3).number_format = '@'
        ws.cell(row=excel_row, column=8).number_format = '@'

        if pd.isna(row['cisloPojistence']):
            ws.cell(row=excel_row, column=1).fill = red
            ws.cell(row=excel_row, column=3).fill = red
            ws.cell(row=excel_row, column=2).fill = red
            ws.cell(row=excel_row, column=4).fill = red
            ws.cell(row=excel_row, column=5).fill = red
        else:
            ws.cell(row=excel_row, column=1).fill = green
            ws.cell(row=excel_row, column=3).fill = green
            jmeno1 = normalize_name(row['jmeno_prijmeni'])
            jmeno2 = normalize_name(row['jmeno_prijmeni.1'])
            if jmeno1 == jmeno2:
                ws.cell(row=excel_row, column=2).fill = green
                ws.cell(row=excel_row, column=4).fill = green
                ws.cell(row=excel_row, column=5).fill = green
            else:
                ws.cell(row=excel_row, column=2).fill = orange
                ws.cell(row=excel_row, column=4).fill = orange
                ws.cell(row=excel_row, column=5).fill = red

        if pd.notna(row['cisloPojistence_navic']):
            ws.cell(row=excel_row, column=8).fill = orange
            ws.cell(row=excel_row, column=9).fill = orange

    ws.auto_filter.ref = ws.dimensions
    wb.save(excel_path)
    print(f"Výsledek uložen do {excel_path}")
    return excel_path

if __name__ == "__main__":
    print("Program začíná...")

    # Vytvoření potřebných složek
    for directory in [DATA_DIR, RESULT_DIR, BACKUP_DIR]:
        if not os.path.exists(directory):
            os.makedirs(directory)
            print(f"Vytvořena složka: {directory}")

    # Příprava databáze
    if not os.path.exists(DB_FILE):
        conn = sqlite3.connect(DB_FILE)
        conn.close()
    if os.path.exists(DB_FILE):
        print(f"Mažu předešlá data v souboru: {DB_FILE}")
        conn = sqlite3.connect(DB_FILE)
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = [row[0] for row in cur.fetchall()]
        for table in tables:
            cur.execute(f"DELETE FROM {table};")
        conn.commit()
        conn.close()

    # Kontrola vstupních souborů
    excel_files = [f for f in os.listdir(DATA_DIR) if f.lower().endswith(('.xlsx', '.xls'))]
    xml_files = [f for f in os.listdir(DATA_DIR) if f.lower().endswith('.xml')]

    if len(excel_files) != 1 or len(xml_files) != 1:
        print("Ve složce /app/data musí být právě jeden Excel soubor (.xlsx/.xls) a pouze jeden XML soubor (.xml).")
        print(f"Excel soubory: {excel_files}")
        print(f"XML soubory: {xml_files}")
        print("Upravte obsah složky a spusťte znovu.")
        sys.exit(1)

    excel_file = os.path.join(DATA_DIR, excel_files[0])
    xml_file = os.path.join(DATA_DIR, xml_files[0])

    # Zpracování dat
    excel_to_sqlite(excel_file, DB_FILE)
    xml_to_sqlite(xml_file, DB_FILE)
    result_file = compare_and_export()

    # Vytvoření zálohy
    backup_dir = create_backup_dir(excel_file, xml_file, DB_FILE)
    if backup_dir:
        print(f"Záloha byla úspěšně vytvořena v: {backup_dir}")
    else:
        print("Varování: Záloha nebyla vytvořena!")